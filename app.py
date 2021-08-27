'''
import ecommerce.shipping
from ecommerce import shipping

ecommerce.shipping.calc_shipping()
shipping.calc_shipping()
'''

import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from openpyxl.xml.constants import MAX_ROW, MIN_COLUMN

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row,3)
        price = cell.value * 0.9
        cell_price = sheet.cell(row,4)
        cell_price.value = price

    values = Reference(sheet,
                        min_row=2,
                        max_row=sheet.max_row,
                        min_col=4,
                        max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,"e2")

    wb.save(filename)


process_workbook("transactions2.xlsx")




'''
clear
& python d:/Solutions/PythonTraining/app.py
'''